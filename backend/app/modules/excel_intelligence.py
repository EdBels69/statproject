"""
ExcelIntelligence — smart parser for messy Excel/CSV files.

Handles: merged cells, multi-row headers, footnote rows, encoding detection,
auto header-row detection, multi-sheet awareness.
"""
from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd


@dataclass
class ExcelStructure:
    """Result of structural analysis of an Excel/CSV file."""
    header_row: int = 0
    data_start_row: int = 1
    skip_footer_rows: int = 0
    sheet_name: Optional[str] = None
    sheet_names: List[str] = field(default_factory=list)
    merged_cells_resolved: int = 0
    multi_header_rows: List[int] = field(default_factory=list)
    encoding: Optional[str] = None
    decoration_rows_skipped: int = 0
    footnote_rows_skipped: int = 0
    issues: List[str] = field(default_factory=list)
    log: List[Dict[str, Any]] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _is_decoration_row(row: pd.Series) -> bool:
    """Check if row is a separator / decoration (all NaN, or all same value)."""
    non_null = row.dropna()
    if non_null.empty:
        return True
    vals = non_null.astype(str).str.strip()
    vals = vals[vals != ""]
    if vals.empty:
        return True
    if vals.nunique() == 1 and len(vals) >= 2:
        v = vals.iloc[0].lower()
        # typical separators: "---", "***", all-whitespace, repeated char
        if len(set(v)) <= 2 and len(v) >= 3:
            return True
    return False


def _is_footnote_like(row: pd.Series, total_cols: int) -> bool:
    """Check if row looks like a footnote/comment (text in first cell, rest empty)."""
    non_null = row.dropna()
    filled = non_null[non_null.astype(str).str.strip() != ""]
    if filled.empty:
        return True
    if len(filled) <= 2 and total_cols >= 3:
        first_val = str(filled.iloc[0]).strip().lower()
        # typical footnotes start with *, #, note, примечание
        if any(first_val.startswith(p) for p in ("*", "#", "note", "примечан", "источник", "source")):
            return True
        # Row with only 1 cell filled out of many columns is likely decoration
        if len(filled) == 1 and total_cols >= 4:
            return True
    return False


def _score_header_row(row: pd.Series) -> float:
    """Score a row on how likely it is to be the header."""
    non_null = row.dropna()
    if non_null.empty:
        return 0.0

    vals = non_null.astype(str).str.strip()
    vals = vals[vals != ""]
    if vals.empty:
        return 0.0

    unique_ratio = vals.nunique() / max(1, len(vals))
    non_numeric = sum(1 for v in vals if not _is_numeric_string(str(v)))
    text_ratio = non_numeric / max(1, len(vals))
    short_text = sum(1 for v in vals if 1 <= len(str(v)) <= 60)
    short_ratio = short_text / max(1, len(vals))

    return (unique_ratio * 0.4) + (text_ratio * 0.35) + (short_ratio * 0.25)


def _is_numeric_string(s: str) -> bool:
    s = s.strip().replace(",", ".").replace(" ", "")
    if s in ("", "-", "+", ".", ","):
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def _detect_encoding(file_path: str) -> str:
    """Detect file encoding using charset_normalizer or chardet."""
    try:
        from charset_normalizer import from_path
        result = from_path(file_path)
        best = result.best()
        if best:
            return best.encoding
    except ImportError:
        pass
    try:
        import chardet
        with open(file_path, "rb") as f:
            raw = f.read(32768)
        det = chardet.detect(raw)
        if det and det.get("encoding"):
            return det["encoding"]
    except ImportError:
        pass
    return "utf-8"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

class ExcelIntelligence:
    """Understands messy Excel files and produces clean DataFrames."""

    def analyze_structure(
        self,
        file_path: str,
        *,
        original_filename: Optional[str] = None,
        sheet_name: Optional[str] = None,
        max_scan_rows: int = 30,
    ) -> ExcelStructure:
        """
        Analyze file structure without loading the full dataset.
        Returns ExcelStructure with detected header row, skip rows, etc.
        """
        result = ExcelStructure()
        ext = os.path.splitext(original_filename or file_path)[1].lower()

        if ext == ".csv":
            result.encoding = _detect_encoding(file_path)
            result.log.append({"action": "detect_encoding", "encoding": result.encoding})

        if ext in (".xlsx", ".xls"):
            result = self._analyze_excel_structure(file_path, result, sheet_name, max_scan_rows)
        elif ext == ".csv":
            result = self._analyze_csv_structure(file_path, result, max_scan_rows)

        return result

    def read_clean(
        self,
        file_path: str,
        *,
        original_filename: Optional[str] = None,
        structure: Optional[ExcelStructure] = None,
    ) -> Tuple[pd.DataFrame, ExcelStructure]:
        """
        Read file into a clean DataFrame using detected or provided structure.
        Handles merged cells, multi-headers, decoration rows, footnotes.
        """
        if structure is None:
            structure = self.analyze_structure(
                file_path, original_filename=original_filename
            )

        ext = os.path.splitext(original_filename or file_path)[1].lower()

        if ext in (".xlsx", ".xls"):
            df = self._read_excel_clean(file_path, structure)
        elif ext == ".csv":
            df = self._read_csv_clean(file_path, structure)
        elif ext == ".parquet":
            df = pd.read_parquet(file_path)
        elif ext == ".json":
            df = pd.read_json(file_path)
        else:
            raise ValueError(f"Unsupported file format: {ext}")

        # Remove fully empty rows and columns
        empty_rows_before = len(df)
        df = df.dropna(how="all")
        empty_rows_removed = empty_rows_before - len(df)
        if empty_rows_removed > 0:
            structure.log.append({
                "action": "remove_empty_rows",
                "count": empty_rows_removed,
            })

        empty_col_positions = []
        for i in range(len(df.columns)):
            try:
                if df.iloc[:, i].isna().all():
                    empty_col_positions.append(i)
            except Exception:
                pass
        if empty_col_positions:
            empty_names = [str(df.columns[i]) for i in empty_col_positions]
            keep_positions = [i for i in range(len(df.columns)) if i not in set(empty_col_positions)]
            df = df.iloc[:, keep_positions]
            structure.log.append({
                "action": "remove_empty_columns",
                "columns": empty_names,
            })

        # Remove decoration rows from the body
        decoration_idx = []
        for idx in df.index:
            if _is_decoration_row(df.loc[idx]):
                decoration_idx.append(idx)
        if decoration_idx:
            df = df.drop(index=decoration_idx)
            structure.decoration_rows_skipped += len(decoration_idx)
            structure.log.append({
                "action": "remove_decoration_rows",
                "count": len(decoration_idx),
            })

        # Remove footnote rows from the tail
        total_cols = len(df.columns)
        footnote_count = 0
        while len(df) > 0 and _is_footnote_like(df.iloc[-1], total_cols):
            df = df.iloc[:-1]
            footnote_count += 1
        if footnote_count:
            structure.footnote_rows_skipped += footnote_count
            structure.log.append({
                "action": "remove_footnote_rows",
                "count": footnote_count,
            })

        df = df.reset_index(drop=True)
        return df, structure

    # ------------------------------------------------------------------
    # Excel-specific
    # ------------------------------------------------------------------

    def _analyze_excel_structure(
        self,
        file_path: str,
        result: ExcelStructure,
        sheet_name: Optional[str],
        max_scan_rows: int,
    ) -> ExcelStructure:
        try:
            xl = pd.ExcelFile(file_path, engine="openpyxl")
            result.sheet_names = xl.sheet_names
            target_sheet = sheet_name or xl.sheet_names[0]
            result.sheet_name = target_sheet
        except Exception as exc:
            result.issues.append(f"Cannot open Excel: {exc}")
            return result

        # Read raw data without header for scanning
        try:
            raw = pd.read_excel(
                file_path,
                sheet_name=target_sheet,
                header=None,
                nrows=max_scan_rows,
                engine="openpyxl",
            )
        except Exception as exc:
            result.issues.append(f"Cannot read sheet '{target_sheet}': {exc}")
            return result

        # Detect merged cells
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb[target_sheet]
            merged_count = len(ws.merged_cells.ranges)
            if merged_count > 0:
                result.merged_cells_resolved = merged_count
                result.log.append({
                    "action": "detect_merged_cells",
                    "count": merged_count,
                })
            wb.close()
        except Exception:
            pass

        # Auto-detect header row
        result = self._detect_header_row(raw, result)

        return result

    def _analyze_csv_structure(
        self,
        file_path: str,
        result: ExcelStructure,
        max_scan_rows: int,
    ) -> ExcelStructure:
        encoding = result.encoding or "utf-8"
        try:
            raw = pd.read_csv(
                file_path,
                header=None,
                nrows=max_scan_rows,
                encoding=encoding,
                on_bad_lines="skip",
            )
        except Exception as exc:
            result.issues.append(f"Cannot read CSV: {exc}")
            return result

        result = self._detect_header_row(raw, result)
        return result

    def _detect_header_row(self, raw: pd.DataFrame, result: ExcelStructure) -> ExcelStructure:
        """Find the most likely header row in the first N rows."""
        if raw.empty:
            return result

        best_score = -1.0
        best_row = 0

        for i in range(min(len(raw), 15)):
            score = _score_header_row(raw.iloc[i])
            if score > best_score:
                best_score = score
                best_row = i

        result.header_row = best_row
        result.data_start_row = best_row + 1

        # Check for multi-row headers (rows above header that also look header-ish)
        if best_row > 0:
            for i in range(best_row):
                sc = _score_header_row(raw.iloc[i])
                if sc > 0.3:
                    result.multi_header_rows.append(i)

        # Special case: merged-cell group headers.
        # When best_row=0 has merged cells (many NaN = col_N later), row 1
        # often contains the real column names (sub-headers).
        # Detect this by scoring the next row and checking NaN density in row 0.
        if best_row == 0 and len(raw) > 1:
            next_score = _score_header_row(raw.iloc[1])
            row0 = raw.iloc[0]
            null_ratio = row0.isna().mean()
            # If next row also looks like a header AND row 0 has many NaN
            # (merged cell groups that expand across multiple columns)
            if next_score > 0.45 and null_ratio > 0.25:
                # Use row 0 as group prefix, row 1 as actual header
                result.multi_header_rows = [0]
                result.header_row = 1
                result.data_start_row = 2
                result.log.append({
                    "action": "detect_merged_group_header",
                    "group_row": 0,
                    "header_row": 1,
                    "group_row_null_ratio": round(float(null_ratio), 3),
                    "next_row_score": round(next_score, 3),
                })

        result.log.append({
            "action": "detect_header_row",
            "header_row": result.header_row,
            "score": round(best_score, 3),
            "multi_header_rows": result.multi_header_rows,
        })

        return result

    # ------------------------------------------------------------------
    # File reading with intelligence
    # ------------------------------------------------------------------

    def _read_excel_clean(self, file_path: str, structure: ExcelStructure) -> pd.DataFrame:
        sheet = structure.sheet_name or 0

        # If multi-header detected, merge header rows
        if structure.multi_header_rows:
            return self._read_excel_multi_header(file_path, structure)

        # Handle merged cells by reading with openpyxl and forward-filling
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb[sheet if isinstance(sheet, str) else wb.sheetnames[sheet]]

            # Unmerge cells and forward-fill values
            for merge_range in list(ws.merged_cells.ranges):
                top_left_value = ws.cell(merge_range.min_row, merge_range.min_col).value
                ws.unmerge_cells(str(merge_range))
                for row in range(merge_range.min_row, merge_range.max_row + 1):
                    for col in range(merge_range.min_col, merge_range.max_col + 1):
                        ws.cell(row=row, column=col, value=top_left_value)
                structure.log.append({
                    "action": "unmerge_cell",
                    "range": str(merge_range),
                    "value": str(top_left_value)[:50] if top_left_value else None,
                })
            wb.close()
        except Exception:
            pass

        df = pd.read_excel(
            file_path,
            sheet_name=sheet,
            header=structure.header_row,
            engine="openpyxl",
        )

        # Clean column names
        df.columns = [
            str(c).strip() if not str(c).startswith("Unnamed") else f"col_{i}"
            for i, c in enumerate(df.columns)
        ]

        return df

    def _read_excel_multi_header(self, file_path: str, structure: ExcelStructure) -> pd.DataFrame:
        """Read Excel with multi-row headers, merging them into single header.

        For merged-cell group headers (row 0 = group names with NaN gaps,
        row 1 = sub-column names), we forward-fill the group row so every
        sub-column inherits its parent group name as a prefix.
        """
        all_header_rows = sorted(set(structure.multi_header_rows + [structure.header_row]))
        sheet = structure.sheet_name or 0

        raw = pd.read_excel(
            file_path,
            sheet_name=sheet,
            header=None,
            engine="openpyxl",
        )

        # Forward-fill each group/prefix row across NaN gaps (merged cell artifact)
        for row_idx in all_header_rows:
            if row_idx < len(raw):
                row_series = raw.iloc[row_idx].copy()
                raw.iloc[row_idx] = row_series.ffill()

        # Build merged column names from header rows
        n_cols = len(raw.columns)
        merged_names: list = []
        seen_names: dict = {}
        for col_idx in range(n_cols):
            parts = []
            for row_idx in all_header_rows:
                if row_idx < len(raw):
                    val = raw.iloc[row_idx, col_idx]
                    if pd.notna(val):
                        s = str(val).strip()
                        if s and s not in parts:
                            parts.append(s)
            name = " / ".join(parts) if parts else f"col_{col_idx}"
            # Make duplicate names unique
            if name in seen_names:
                seen_names[name] += 1
                name = f"{name}_{seen_names[name]}"
            else:
                seen_names[name] = 0
            merged_names.append(name)

        # Data starts after the last header row
        data_start = max(all_header_rows) + 1
        df = raw.iloc[data_start:].copy()
        df.columns = merged_names
        df = df.reset_index(drop=True)

        structure.log.append({
            "action": "merge_multi_header",
            "header_rows": all_header_rows,
            "merged_columns": merged_names[:8],
        })

        return df

    def _read_csv_clean(self, file_path: str, structure: ExcelStructure) -> pd.DataFrame:
        encoding = structure.encoding or "utf-8"
        df = pd.read_csv(
            file_path,
            header=structure.header_row,
            encoding=encoding,
            on_bad_lines="skip",
        )

        # Clean column names
        df.columns = [
            str(c).strip() if not str(c).startswith("Unnamed") else f"col_{i}"
            for i, c in enumerate(df.columns)
        ]

        return df
